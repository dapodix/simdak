from __future__ import annotations
import os
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import Union
from . import RkasData, SimdakPaud
from simdak.template import TEMPLATE_FILE

CWD = os.getcwd()


def exports(email: str, password: str, filename: str = "", sheet: str = "Sheet1"):
    print("Export")
    simdak = SimdakPaud(email, password)
    rkas = simdak.rkas()[0]
    rkas_datas = rkas.get(save_as=RkasData)
    filename = filename or f"{rkas.npsn}.xlsx"
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"
    filepath = os.path.join(CWD, filename)
    ws: Worksheet = None
    wb = load_workbook(TEMPLATE_FILE)
    ws = wb.active
    ws.title = sheet
    ws["A1"] = "No"
    for k, v in RkasData.MAPPING.items():
        ws[f"{v}1"] = k.replace("_", " ").title()
    for i, rkas_data in enumerate(rkas_datas):
        rkas_data.to_row(ws, i + 2)
        ws[f"A{i+2}"] = i + 1
    wb.save(filepath)


def imports(
    email: str,
    password: str,
    filename: str = "",
    start: int = 1,
    sheet: str = "Sheet1",
    header: bool = True,
):
    print("Import")
    simdak = SimdakPaud(email, password)
    rkas = simdak.rkas()[0]
    filepath = os.path.join(CWD, filename)
    ws: Worksheet = None
    if os.path.isfile(filepath):
        wb = load_workbook(filepath)
        sheets = wb.get_sheet_names()
        ws = wb.get_sheet_by_name(sheet) if sheet in sheets else wb.active
    row = start + 1 if header else start
    while True:
        if not ws[f"B{row}"]:
            break
        data = RkasData.from_row(ws, row)
        result = rkas.create(data)
        if result:
            result.to_row(ws, row)
